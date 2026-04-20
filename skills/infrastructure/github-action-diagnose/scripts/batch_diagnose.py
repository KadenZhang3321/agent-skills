#!/usr/bin/env python3
"""
batch_diagnose.py - 批量诊断 xlsx 文件中的 Job URL，结果写回原文件

用法:
  python batch_diagnose.py --input Fail_CI_Problem/failed-runs-2026-04-17.xlsx
  python batch_diagnose.py --input Fail_CI_Problem/failed-runs-2026-04-17.xlsx --api-key sk-xxx

特性:
  - 串行逐条诊断（简单可靠）
  - 断点续传（跳过已诊断的行）
  - 每完成一条立即保存（防止中断丢失）
  - 进度显示（[3/20] 诊断中...）

依赖:
  pip install openpyxl openai
  gh CLI 已安装并认证
"""

import argparse
import json
import os
import re
import subprocess
import sys
import time
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPT_DIR))

# Force UTF-8 for subprocess
os.environ["PYTHONIOENCODING"] = "utf-8"

from diagnose_job import (
    SKILL_DIR,
    TOOLS,
    load_skill_content,
    parse_job_url,
    execute_tool,
    call_qwen,
)

# 导入 MCP 工具
from diagnose_job import (
    mcp_list_pods_tool,
    mcp_export_logs_tool,
    mcp_get_export_tool,
    mcp_get_runner_logs_tool,
)

RESULT_COLUMN = "诊断结果"


def parse_args():
    parser = argparse.ArgumentParser(description="批量诊断 xlsx 中的 Job URL")
    parser.add_argument(
        "--input",
        required=True,
        help="输入的 xlsx 文件路径",
    )
    parser.add_argument(
        "--api-key",
        default=os.environ.get("DASHSCOPE_API_KEY", ""),
        help="百炼 API Key",
    )
    parser.add_argument(
        "--model",
        default="qwen3.6-plus",
        help="模型名称（默认 qwen3.6-plus）",
    )
    parser.add_argument(
        "--max-turns",
        type=int,
        default=15,
        help="每条最大工具调用轮数",
    )
    parser.add_argument(
        "--skill-dir",
        default=str(SKILL_DIR),
        help="Skill 目录路径",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="只诊断前 N 条（0 表示全部）",
    )
    return parser.parse_args()


def build_system_prompt(skill_dir: str, gh_path: str) -> str:
    skill_content = load_skill_content(skill_dir)

    env_info = f"""环境信息：
- 操作系统: Windows (PowerShell 5.1)
- gh CLI 路径: {gh_path or "C:\\Program Files\\GitHub CLI\\gh.exe"}
- 没有 bash、python3、curl 等 Unix 工具
- 使用 cmd 命令或 PowerShell 命令
- 使用 findstr 替代 grep
- 使用 type 替代 cat
- 使用 dir 替代 ls"""

    return f"""你是一个 CI 故障诊断专家。请严格按照以下 SKILL 文档的指引，诊断用户提供的失败 Job。

<skill_document>
{skill_content}
</skill_document>

{env_info}

## 推荐诊断流程

**第一步（必须）：调用 fetch_run_script**
- `fetch_run_script(repo, job_id)` 会优先运行 `fetch-run.sh`（bash），失败则回退到 `fetch_run.py`
- 一次性输出：
  - Job 基本信息、Runner、失败步骤
  - PR 上下文和变更文件
  - Annotations（通常直接揭示根因）
  - 预过滤的错误日志（如需要）
- **这个工具的输出通常已足够完成诊断**

**第二步（按需）：当 GitHub 日志不足以定位根因时，使用 MCP 查询集群日志**
- 场景：GitHub 日志只显示 exit code 1，但不知道具体原因
- `mcp_get_runner_logs(runner_name, start, end, keywords)` - 一键获取 Runner Pod 日志
  - runner_name: 从 fetch_run_script 输出中获取
  - start/end: 从 fetch_run_script 输出的时间范围获取，格式 YYYY-MM-DD HH:MM:SS
  - **keywords: 必须传，根据错误类型选择过滤词**：
    - 超时/网络: `"ETIMEDOUT|timeout|connection refused"`
    - 内存: `"OOM|OutOfMemory|Killed"`
    - 运行时错误: `"RuntimeError|AssertionError|EngineDeadError"`
    - 编译错误: `"error:|failed|fatal"`
    - 不确定: `"Error|Failed|exit code|Exception"`

**重要规则：**
1. **必须先调用 fetch_run_script**，它的输出通常足以完成诊断
2. **fetch_run_script 输出后，如果信息足够，直接输出诊断报告，不要再调用其他工具**
3. 只有在 GitHub 日志不足以定位根因时，才使用 MCP 查询集群日志
4. 最终诊断结果直接输出，不要调用多余的工具
5. 不粘贴大段原始日志，只引用最关键的一行错误标识
6. 每个问题标注责任方（基础设施团队 / PR 作者）

Skill 目录: {skill_dir}
当前工作目录: {os.getcwd()}"""


def _find_gh() -> str:
    try:
        where = subprocess.run(
            "where gh.exe",
            shell=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        if where.stdout.strip():
            return where.stdout.strip().split("\n")[0]
    except Exception:
        pass
    return "C:\\Program Files\\GitHub CLI\\gh.exe"


def diagnose_single(
    job_url: str,
    system_prompt: str,
    api_key: str,
    model: str,
    max_turns: int,
) -> tuple[str, dict]:
    """诊断单个 Job URL，返回 (诊断报告文本, token用量)"""
    repo, job_id = parse_job_url(job_url)
    total_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}

    messages = [
        {"role": "system", "content": system_prompt},
        {
            "role": "user",
            "content": f"请诊断这个失败的 Job: {job_url}\n\n仓库: {repo}\nJob ID: {job_id}",
        },
    ]

    for turn in range(1, max_turns + 1):
        message = call_qwen(messages, model, api_key, TOOLS)

        usage = message.get("usage", {})
        total_usage["prompt_tokens"] += usage.get("prompt_tokens", 0)
        total_usage["completion_tokens"] += usage.get("completion_tokens", 0)
        total_usage["total_tokens"] += usage.get("total_tokens", 0)

        if message.get("tool_calls"):
            tool_calls = message["tool_calls"]
            tool_results = []

            for tc in tool_calls:
                tool_name = tc["function"]["name"]
                tool_args = json.loads(tc["function"]["arguments"])
                result = execute_tool(tool_name, tool_args)

                tool_results.append(
                    {
                        "role": "tool",
                        "tool_call_id": tc["id"],
                        "content": result,
                    }
                )

            messages.append(message)
            messages.extend(tool_results)

        else:
            return message.get("content", "无输出"), total_usage

    return f"（达到最大轮数限制 {max_turns}，诊断未完成）", total_usage


def main():
    args = parse_args()

    if not args.api_key:
        print("错误: 未设置 API Key")
        print("请设置环境变量 DASHSCOPE_API_KEY 或使用 --api-key 参数")
        sys.exit(1)

    xlsx_path = Path(args.input)
    if not xlsx_path.exists():
        print(f"错误: 文件不存在 {xlsx_path}")
        sys.exit(1)

    gh_path = _find_gh()
    system_prompt = build_system_prompt(args.skill_dir, gh_path)

    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if RESULT_COLUMN not in headers:
        headers.append(RESULT_COLUMN)
        ws.cell(row=1, column=len(headers), value=RESULT_COLUMN)

    url_col_idx = headers.index("Job URL") + 1
    result_col_idx = headers.index(RESULT_COLUMN) + 1

    total_rows = ws.max_row - 1
    pending_urls = []

    for row in range(2, ws.max_row + 1):
        url = ws.cell(row=row, column=url_col_idx).value
        result = ws.cell(row=row, column=result_col_idx).value
        if not url:
            continue
        if result and str(result).strip():
            print(f"[跳过] 第 {row - 1} 行: {url[:60]}... (已有诊断结果)")
        else:
            pending_urls.append((row, url))

    if args.limit > 0:
        pending_urls = pending_urls[: args.limit]
        print(f"（限制前 {args.limit} 条）\n")

    if not pending_urls:
        print("\n所有行已诊断完毕，无需处理。")
        return

    print(f"\n共 {len(pending_urls)} 条待诊断，总计 {total_rows} 行\n")

    success_count = 0
    fail_count = 0
    grand_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}

    for i, (row, url) in enumerate(pending_urls, 1):
        print(f"{'=' * 50}")
        print(f"[{i}/{len(pending_urls)}] 诊断: {url}")
        print(f"{'=' * 50}")

        try:
            start = time.time()
            result, usage = diagnose_single(
                job_url=url,
                system_prompt=system_prompt,
                api_key=args.api_key,
                model=args.model,
                max_turns=args.max_turns,
            )
            elapsed = time.time() - start

            grand_usage["prompt_tokens"] += usage["prompt_tokens"]
            grand_usage["completion_tokens"] += usage["completion_tokens"]
            grand_usage["total_tokens"] += usage["total_tokens"]

            ws.cell(row=row, column=result_col_idx, value=result)
            wb.save(str(xlsx_path))

            print(
                f"\n✅ 完成 ({elapsed:.0f}s) | Token: {usage['total_tokens']:,} (输入 {usage['prompt_tokens']:,} / 输出 {usage['completion_tokens']:,})"
            )
            success_count += 1

        except Exception as e:
            error_msg = f"诊断失败: {str(e)}"
            ws.cell(row=row, column=result_col_idx, value=error_msg)
            wb.save(str(xlsx_path))

            print(f"\n❌ {error_msg}")
            fail_count += 1

        print()

    # 计算费用 (Qwen3.6-Plus 中国内地: 输入 2元/百万, 输出 12元/百万)
    input_cost = grand_usage["prompt_tokens"] / 1_000_000 * 2
    output_cost = grand_usage["completion_tokens"] / 1_000_000 * 12
    total_cost = input_cost + output_cost

    print(f"\n{'=' * 50}")
    print(f"批量诊断完成")
    print(f"  成功: {success_count}")
    print(f"  失败: {fail_count}")
    print(f"  Token 用量: {grand_usage['total_tokens']:,}")
    print(f"    输入: {grand_usage['prompt_tokens']:,}")
    print(f"    输出: {grand_usage['completion_tokens']:,}")
    print(
        f"  预估费用: ¥{total_cost:.4f} (输入 ¥{input_cost:.4f} + 输出 ¥{output_cost:.4f})"
    )
    print(f"  输出: {xlsx_path}")
    print(f"{'=' * 50}")


if __name__ == "__main__":
    main()
